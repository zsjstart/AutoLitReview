"""
Research idea -> multi-concept OpenAlex search -> gpt-oss analysis -> grouped Excel.

Pipeline:
  1. Concept extraction : LLM generates N concepts from the idea (no selection pass).
  2. Collection         : OpenAlex search per concept, year-filtered AND optionally
                          restricted to a research field (--domain), pooled + deduped by DOI.
  3. Phase 1 (local)    : chunked LLM calls extract per-paper summary + target object.
  4. Relevance gate     : one cheap LLM call drops papers not genuinely about the idea.
  5. Phase 2 (global)   : one LLM call discovers categories across the survivors, then assigns.
  6. Excel              : Papers sheet (Title, Year, Summary, Target Object, Category, Source)
                          + a discovered-Categories sheet.

Requirements:
  pip install openai requests openpyxl
  export OPENALEX_API_KEY=...        (free key at openalex.org/settings/api)
  a vLLM server running gpt-oss-120b on an OpenAI-compatible endpoint.

Usage:
  python llm_papers_collector_and_processor.py "your research idea" --concepts 4
  python llm_papers_collector_and_processor.py "your idea" --domain cybersecurity \
         --concepts 4 --per-concept 7 --year-from 2020 --year-to 2026
  python llm_papers_collector_and_processor.py "your idea" \
         --manual-concepts "OSINT" "attribute inference" "de-anonymization"
"""
import os
import json
import argparse
import requests
from openai import OpenAI

# ----- LLM provider configuration (set by configure_provider via --provider) -----
# All four providers are reached through the OpenAI-compatible SDK shape; they differ
# only in base_url, key, default model, and whether text arrives in reasoning_content.
client = None
MODEL = None
HAS_REASONING_CONTENT = False   # gpt-oss/vLLM puts text in reasoning_content; others don't

PROVIDER_DEFAULTS = {
    # provider:  (base_url, key_env, default_model, has_reasoning_content)
    "vllm":   ("http://localhost:8000/v1",                              None,                "openai/gpt-oss-120b", True),
    "openai": ("https://api.openai.com/v1",                             "OPENAI_API_KEY",    "gpt-4o",              False),
    "gemini": ("https://generativelanguage.googleapis.com/v1beta/openai/", "GEMINI_API_KEY", "gemini-3.5-flash",    False),
    "claude": ("https://api.anthropic.com/v1/",                         "ANTHROPIC_API_KEY", "claude-sonnet-4-6",   False),
}

def configure_provider(provider, model_override=None):
    global client, MODEL, HAS_REASONING_CONTENT
    provider = (provider or "vllm").lower()
    if provider not in PROVIDER_DEFAULTS:
        raise SystemExit(f"Unknown provider '{provider}'. "
                         f"Choose from: {', '.join(PROVIDER_DEFAULTS)}")
    base_url, key_env, default_model, has_rc = PROVIDER_DEFAULTS[provider]

    if key_env:
        key = os.environ.get(key_env, "")
        if not key:
            raise SystemExit(f"Set {key_env} for provider '{provider}'.")
    else:
        key = "EMPTY"   # local vLLM needs no real key

    if provider == "vllm":
        base_url = os.environ.get("LLM_BASE_URL", base_url)

    client = OpenAI(base_url=base_url, api_key=key)
    MODEL = model_override or os.environ.get("LLM_MODEL") or default_model
    HAS_REASONING_CONTENT = has_rc
    print(f"Provider: {provider} (model {MODEL})")

# ----- OpenAlex (API key required since 2026-02-13) -----
OPENALEX_KEY = os.environ.get("OPENALEX_API_KEY", "")

# ----- Internal constant: token-safety chunk size for Phase 1 -----
ANALYSIS_BATCH = 5

# ----- OpenAlex field IDs (level-2 of the Domain->Field->Subfield->Topic hierarchy) -----
# Used by --domain as a HARD server-side filter (primary_topic.field.id).
DOMAIN_FIELDS = {
    "cybersecurity": 17, "computer science": 17, "cs": 17,
    "engineering": 22, "materials science": 25, "mathematics": 26,
    "physics": 31, "biology": 13, "chemistry": 16,
    "medicine": 27, "psychology": 32, "social sciences": 33,
    "economics": 20, "neuroscience": 28, "environmental science": 23,
}


# ============================================================
# LLM helpers
# ============================================================
def llm(prompt, temperature=0.0, max_tokens=4000):
    """Single chat completion across providers (OpenAI-compatible shape)."""
    r = client.chat.completions.create(
        model=MODEL,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=max_tokens,
        temperature=temperature,
    )
    choice = r.choices[0]
    content = getattr(choice.message, "content", None)
    if not content and HAS_REASONING_CONTENT:
        content = getattr(choice.message, "reasoning_content", None)
    if not content:
        raise RuntimeError(
            f"Empty model response (finish_reason={choice.finish_reason}). "
            f"Raise max_tokens for this call."
        )
    return content.strip()


def llm_json(prompt, temperature=0.0, max_tokens=4000):
    """Call the LLM and parse JSON, with one automatic retry (more tokens) on failure."""
    def _parse(text):
        text = text.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        starts = [text.find(c) for c in "{[" if text.find(c) != -1]
        start = min(starts) if starts else 0
        end = max(text.rfind("}"), text.rfind("]")) + 1
        return json.loads(text[start:end])

    try:
        return _parse(llm(prompt, temperature=temperature, max_tokens=max_tokens))
    except (json.JSONDecodeError, ValueError, RuntimeError):
        raw = llm(prompt + "\n\nIMPORTANT: Return ONLY valid JSON, no other text.", temperature=0.0, max_tokens=max_tokens + 4000)
        return _parse(raw)


def _coerce_list(obj):
    """Accept ['a','b'] or {'concepts':[...]} and return a flat list of strings."""
    if isinstance(obj, dict):
        for v in obj.values():
            if isinstance(v, list):
                obj = v
                break
        else:
            obj = []
    return [str(x).strip() for x in obj if str(x).strip()]


# ============================================================
# Step 1: concept extraction (generate N; no selection pass)
# ============================================================
def extract_concepts(idea, n):
    gen_prompt = f"""Given this research idea:
{idea}

Generate {n} literature-search concepts for finding papers ON THIS SPECIFIC TOPIC.

Requirements:
- Use terminology commonly found in paper titles, abstracts,
  author keywords, or research taxonomies.
- Prefer established academic concepts.
- Avoid synthesized descriptive phrases.
- Avoid combining multiple concepts into one phrase.
- Each concept should be searchable independently in DBLP/OpenAlex.


Return ONLY a JSON array of strings."""
    candidates = _coerce_list(llm_json(gen_prompt, temperature=0.2, max_tokens=2000))
    return candidates[:n]


# ============================================================
# Step 2: OpenAlex collection (with optional hard field filter)
# ============================================================
def _reconstruct_abstract(inverted_index):
    if not inverted_index:
        return ""
    words = {}
    for word, positions in inverted_index.items():
        for pos in positions:
            words[pos] = word
    return " ".join(words[i] for i in sorted(words))


def search_concept(concept, per_concept, year_from, year_to, field_id=None,
                   core_only=False):
    filters = ["has_abstract:true", f"publication_year:{year_from}-{year_to}"]
    if field_id:
        filters.append(f"primary_topic.field.id:{field_id}")          # hard domain filter
    if core_only:
        filters.append("primary_location.source.is_core:true")        # reputable venues only
    params = {
        "search": concept,
        "per-page": per_concept,
        "filter": ",".join(filters),
        "sort": "relevance_score:desc",
        "api_key": OPENALEX_KEY,
    }
    resp = requests.get("https://api.openalex.org/works", params=params, timeout=30)
    resp.raise_for_status()
    out = []
    for w in resp.json().get("results", [])[:per_concept]:
        loc = w.get("primary_location") or {}
        source_obj = loc.get("source") or {}
        src = (w.get("doi") or loc.get("landing_page_url")
               or source_obj.get("homepage_url") or "")
        out.append({
            "doi": (w.get("doi") or "").lower(),
            "title": w.get("title") or "",
            "year": w.get("publication_year") or "",
            "abstract": _reconstruct_abstract(w.get("abstract_inverted_index")),
            "source": src,
            "venue": source_obj.get("display_name") or "",
            "is_core": bool(source_obj.get("is_core")),
            "is_doaj": bool(source_obj.get("is_in_doaj")),
        })
    return out


def collect(concepts, per_concept, year_from, year_to, field_id=None, core_only=False):
    seen, papers = set(), []
    for c in concepts:
        print(f"  searching: {c}")
        try:
            results = search_concept(c, per_concept, year_from, year_to, field_id,
                                     core_only)
        except requests.HTTPError as e:
            print(f"    (search failed for '{c}': {e})")
            continue
        for p in results:
            key = p["doi"] or p["title"].strip().lower()
            if key and key not in seen:
                seen.add(key)
                papers.append(p)
    return papers


# ============================================================
# Step 3: Phase 1 - chunked local analysis (summary + target object)
# ============================================================
def analyze_local(papers, batch=ANALYSIS_BATCH):
    results = {}
    for start in range(0, len(papers), batch):
        chunk = papers[start:start + batch]
        block = "\n\n".join(
            f'{start + i}. Title: {p["title"]}\n   Abstract: {p["abstract"]}'
            for i, p in enumerate(chunk)
        )
        prompt = f"""For each paper below, extract two fields.

Output ONLY a JSON object keyed by the paper's index number, e.g.:
{{
  "{start}": {{
    "summary": "one-sentence summary of the paper",
    "target_object": "what the method targets or acts upon"
  }}
}}

Papers:
{block}"""
        part = llm_json(prompt, temperature=0.0, max_tokens=4000)
        if isinstance(part, dict):
            results.update(part)
    return results


# ============================================================
# Step 4: Relevance gate - drop papers not about the idea (summaries only)
# ============================================================
def relevance_filter(idea, papers, local):
    block = "\n".join(
        f'{i}. {p["title"]} - {local.get(str(i), {}).get("summary", "")}'
        for i, p in enumerate(papers)
    )
    prompt = f"""Research idea:
{idea}

Below are candidate papers (index, title, one-line summary). Some were retrieved by
keyword search but may NOT actually be about this research idea.

For EACH paper decide keep or drop. Output ONLY a JSON object mapping each index to
true (keep) or false (drop), e.g.: {{"0": true, "1": false, "2": true}}

Papers:
{block}"""
    verdicts = llm_json(prompt, temperature=0.0, max_tokens=3000)

    keep_idx = set()
    if isinstance(verdicts, dict):
        if any(str(k).isdigit() for k in verdicts.keys()):
            keep_idx = {int(k) for k, v in verdicts.items()
                        if str(k).isdigit() and v in (True, "true", "True")}
        else:
            for v in verdicts.values():
                if isinstance(v, list):
                    keep_idx = {int(i) for i in v if str(i).isdigit()}
                    break
    elif isinstance(verdicts, list):
        keep_idx = {int(i) for i in verdicts if str(i).isdigit()}

    kept_papers, kept_local = [], {}
    for new_i, old_i in enumerate(sorted(keep_idx)):
        if old_i < len(papers):
            kept_papers.append(papers[old_i])
            kept_local[str(new_i)] = local.get(str(old_i), {})
    return kept_papers, kept_local


# ============================================================
# Step 5: Phase 2 - global taxonomy discovery + assignment (summaries only)
# ============================================================
def categorize_global(papers, local):
    block = "\n".join(
        f'{i}. {p["title"]} - {local.get(str(i), {}).get("summary", "")}'
        for i, p in enumerate(papers)
    )
    prompt = f"""Below are {len(papers)} research papers (title + one-line summary).

Do the following IN ORDER:
1. FIRST read the whole set and DISCOVER the natural categories that group them.
   Define the categories yourself; do not use a predefined list. Name each one.
2. THEN assign every paper to one or more of the categories you defined.

Output ONLY a JSON object shaped exactly like:
{{
  "categories": {{"<category name>": "<one-line description>"}},
  "assignments": {{"0": ["<category name>"], "1": ["<category name>", "..."]}}
}}

Papers:
{block}"""
    result = llm_json(prompt, temperature=0.0, max_tokens=12000)
    result.setdefault("categories", {})
    result.setdefault("assignments", {})
    return result


# ============================================================
# Step 6: Excel output
# ============================================================
def write_excel(papers, local, cats, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="4472C4")
    body_font = Font(name="Arial")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"
    headers = ["Title", "Year", "Summary", "Target Object", "Category",
               "Venue", "Venue Quality", "Source"]
    ws.append(headers)
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    assignments = cats.get("assignments", {})
    for i, p in enumerate(papers):
        lo = local.get(str(i), {})
        cat = assignments.get(str(i), [])
        cat_str = ", ".join(cat) if isinstance(cat, list) else str(cat)
        # venue-quality tag from OpenAlex signals
        if p.get("is_core"):
            quality = "Core"
        elif p.get("is_doaj"):
            quality = "DOAJ"
        else:
            quality = "Other"
        ws.append([
            p["title"], str(p["year"]),
            lo.get("summary", ""), lo.get("target_object", ""),
            cat_str, p.get("venue", ""), quality, p["source"],
        ])
    for col, width in zip("ABCDEFGH", [38, 7, 52, 26, 28, 26, 12, 30]):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = wrap
            c.font = body_font
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Categories")
    ws2.append(["Category", "Description"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
    for name, desc in cats.get("categories", {}).items():
        ws2.append([name, desc])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 70
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.alignment = wrap
            c.font = body_font

    wb.save(out)

def filter_redundancy(papers):
    """
    Remove papers with duplicate titles.
    Keeps the first occurrence.
    """
    seen = set()
    filtered = []

    for p in papers:
        title = p.get("title", "").strip().lower()

        if title in seen:
            continue

        seen.add(title)
        filtered.append(p)

    removed = len(papers) - len(filtered)
    print(f"Removed {removed} duplicate papers by title")

    return filtered


# ============================================================
# Orchestration
# ============================================================
def run(idea, concepts_n=4, per_concept=5, year_from=2024, year_to=2026,
        domain=None, core_only=False, manual_concepts=None, out="papers_grouped.xlsx"):
    idea = idea.strip()
    if year_from > year_to:
        print(f"year-from ({year_from}) is after year-to ({year_to}); swapping.")
        year_from, year_to = year_to, year_from

    field_id = None
    if domain:
        field_id = DOMAIN_FIELDS.get(domain.lower().strip())
        if field_id is None:
            print(f"Unknown domain '{domain}' - searching without a field filter. "
                  f"Known domains: {', '.join(sorted(DOMAIN_FIELDS))}")
        else:
            print(f"Domain filter: {domain} (OpenAlex field id {field_id})")
    if core_only:
        print("Venue filter: core sources only (excludes preprints like arXiv)")

    concepts = manual_concepts or extract_concepts(idea, concepts_n)
    if not concepts:
        print("No concepts produced - check the model is responding.")
        return
    print(f"Concepts: {concepts}")

    papers = collect(concepts, per_concept, year_from, year_to, field_id, core_only)
    print(f"Collected {len(papers)} unique papers")
    if not papers:
        print("No papers found. Widen --year-from, broaden the idea, relax --domain, "
              "drop --core-only, or check the OpenAlex key.")
        return
    
    papers = filter_redundancy(papers)
    
    local = analyze_local(papers)
    print("Phase 1 (per-paper analysis) done")
    
    

    papers, local = relevance_filter(idea, papers, local)
    print(f"Relevance gate: {len(papers)} papers kept as on-topic")
    if not papers:
        print("No papers survived the relevance filter. Try broader concepts, a wider "
              "year range, or a higher --per-concept so more candidates are collected.")
        return
    if len(papers) < 4:
        print(f"Warning: only {len(papers)} on-topic papers - grouping will be weak. "
              f"Consider raising --per-concept or widening --year-from.")

    cats = categorize_global(papers, local)
    print(f"Phase 2 (grouping) done - {len(cats.get('categories', {}))} categories discovered")

    write_excel(papers, local, cats, out)
    print(f"Done -> {out}")


def parse_args():
    p = argparse.ArgumentParser(
        description="Research idea -> OpenAlex search -> gpt-oss analysis -> grouped Excel."
    )
    p.add_argument("idea", help="the research idea (in quotes)")
    p.add_argument("--concepts", type=int, default=4,
                   help="number of concepts to generate and search (default 4)")
    p.add_argument("--per-concept", type=int, default=5,
                   help="papers fetched per concept (default 5)")
    p.add_argument("--year-from", type=int, default=2024,
                   help="earliest publication year (default 2024)")
    p.add_argument("--year-to", type=int, default=2026,
                   help="latest publication year (default 2026)")
    p.add_argument("--domain", default=None,
                   help="restrict results to an OpenAlex field for higher precision, "
                        "e.g. 'cybersecurity' (= Computer Science). Known: "
                        + ", ".join(sorted(DOMAIN_FIELDS)))
    p.add_argument("--core-only", action="store_true",
                   help="keep only papers from 'core' (reputable) venues. NOTE: this "
                        "excludes preprint servers like arXiv, so recent preprints are lost.")
    p.add_argument("--provider", default="vllm",
                   choices=["vllm", "openai", "gemini", "claude"],
                   help="LLM backend (default vllm = local gpt-oss). openai/gemini/claude "
                        "need OPENAI_API_KEY / GEMINI_API_KEY / ANTHROPIC_API_KEY respectively.")
    p.add_argument("--model", default=None,
                   help="override the model name for the chosen provider "
                        "(else a sensible per-provider default is used)")
    p.add_argument("--manual-concepts", nargs="+", default=None,
                   help="supply your own concepts, bypassing auto-extraction")
    p.add_argument("--out", default="papers_grouped.xlsx",
                   help="output Excel filename (default papers_grouped.xlsx)")
    return p.parse_args()


if __name__ == "__main__":
    if not OPENALEX_KEY:
        raise SystemExit("Set OPENALEX_API_KEY (free key at openalex.org/settings/api)")
    a = parse_args()
    configure_provider(a.provider, a.model)
    run(a.idea, concepts_n=a.concepts, per_concept=a.per_concept,
        year_from=a.year_from, year_to=a.year_to, domain=a.domain,
        core_only=a.core_only, manual_concepts=a.manual_concepts, out=a.out)
